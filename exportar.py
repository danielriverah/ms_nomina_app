"""
Generador de exportación Excel - formato NOMINA TOPES para Nomipaq
"""
import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


AZUL     = "1F4E79"
AZUL_CLR = "BDD7EE"
VERDE    = "375623"
VERDE_CLR= "E2EFDA"
GRIS     = "808080"
GRIS_CLR = "F2F2F2"
ROJO     = "C00000"
BLANCO   = "FFFFFF"
AMARILLO = "FFFF00"

_thin = Side(style="thin", color="AAAAAA")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _hdr(ws, row, col, val, bg=AZUL_CLR, bold=False, wrap=True, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, size=8, color="000000")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = _border
    if fmt:
        c.number_format = fmt
    return c

def _cell(ws, row, col, val, fmt=None, bold=False, color=None, bg=None, align="right"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=8, color=color or "000000")
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def generar_nomina_topes(periodo, incidencias_list, tipo_trabajador="PERMANENTE") -> bytes:
    """
    Genera el Excel de nómina en formato compatible con Nomipaq.
    Incluye hoja NOMINA (fiscal) y hoja DIFERENCIAS (no fiscal).
    """
    wb = openpyxl.Workbook()

    # ── Hoja 1: NOMINA FISCAL ──────────────────────────────────────────
    ws = wb.active
    ws.title = "NOMINA"

    # Fila 1: Encabezado empresa
    ws.merge_cells("A1:AM1")
    c = ws["A1"]
    c.value = f"MEGA FRESH PRODUCE S. DE R.L. | RFC: MFP050802NS4 | CELAYA, GTO"
    c.font = Font(bold=True, size=10, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # Fila 2: Info del período
    ws.merge_cells("A2:AM2")
    c = ws["A2"]
    c.value = (
        f"NÓMINA SEMANAL {tipo_trabajador} | "
        f"SEMANA {periodo['num_semana']} / {periodo['anio']} | "
        f"PERÍODO: {periodo['fecha_inicio']} AL {periodo['fecha_fin']} | "
        f"FECHA DE PAGO: {periodo['fecha_pago']} | "
        f"UMA: ${periodo['uma_vigente']:.2f} | SM: ${periodo['sm_vigente']:.2f}"
    )
    c.font = Font(bold=True, size=8, color="000000")
    c.fill = PatternFill("solid", fgColor=AZUL_CLR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    # Fila 3: Headers de columnas (formato NOMINA TOPES)
    HEADERS = [
        ("AÑO", 5), ("MES", 4), ("TIPO", 5), ("FECHA\nPAGO", 12),
        ("CLAVE\nTRAB", 7), ("NSS", 14), ("NOMBRE", 30), ("TIPO\nTRAB", 7),
        ("PUESTO", 18), ("SBC", 9), ("DIAS", 5), ("INCAP", 5),
        ("DIAS\nNETOS", 6), ("DIAS\nNOM", 6), ("SD", 9),
        ("SUELDO", 10), ("DESPENSA", 9), ("ASIST.", 9), ("PUNTUAL.", 9),
        ("VACACIONES", 10), ("HRS\nEXTRAS", 9), ("PRIMA\nVAC", 9),
        ("COMPENS.", 9), ("SUMA\nREMUN", 10),
        ("CUOTA\nOBRERA", 9), ("DESC\nALIMENT", 9), ("IMPUESTO", 9),
        ("ISR\nCALCULA", 9), ("SUB EMP\nACRE", 9), ("ISR\nNETO", 9),
        ("SUBEMP\nNET", 9), ("CRED\nINFONAV", 9), ("REDONDEO", 8),
        ("SUMA\nDEDUC", 10), ("NETO\nNUEVO", 10), ("PAGADO", 10),
        ("TIPO\nDIF", 7), ("EXENTO\nHRS EXT", 9), ("NO\nHRS", 6),
        ("EXENTO\nPRIMA VAC", 9),
    ]

    for col_idx, (hdr_name, col_w) in enumerate(HEADERS, start=1):
        _hdr(ws, 3, col_idx, hdr_name, bg=AZUL_CLR, bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws.row_dimensions[3].height = 28

    # Datos
    row = 4
    totales = {k: 0.0 for k in [
        "sueldo","despensa","asist","puntual","vac","hrs_ext","prima_vac","compens",
        "suma_remun","cuota_obrera","isr_calcula","sub_emp","isr_neto","infonavit",
        "suma_deduc","neto","pagado"
    ]}
    num_trab = 0

    for inc in incidencias_list:
        t = inc["trabajador"]
        r = inc["resultado"]
        if not r:
            continue

        # Determinar redondeo
        redondeo = round(r["neto_fiscal"] - int(r["neto_fiscal"] * 100) / 100, 4)
        pagado   = r["neto_fiscal"]

        col = 1
        _cell(ws, row, col, periodo["anio"], fmt="0", align="center"); col+=1
        _cell(ws, row, col, int(periodo["fecha_pago"][5:7]), fmt="0", align="center"); col+=1
        _cell(ws, row, col, "Q", align="center"); col+=1
        _cell(ws, row, col, periodo["fecha_pago"], fmt="DD/MM/YYYY", align="center"); col+=1
        _cell(ws, row, col, t["id"], fmt="0", align="center"); col+=1
        _cell(ws, row, col, t["imss"], align="left"); col+=1
        _cell(ws, row, col, t["nombre_completo"], align="left"); col+=1
        _cell(ws, row, col, "P" if t["tipo_trabajador"] == "PERMANENTE" else "E", align="center"); col+=1
        _cell(ws, row, col, t["puesto"], align="left"); col+=1
        _cell(ws, row, col, t["sbc_dia"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, inc["dias_trabajados"], fmt="0", align="center"); col+=1
        _cell(ws, row, col, inc["dias_incapacidad"], fmt="0", align="center"); col+=1
        _cell(ws, row, col, inc["dias_trabajados"] - inc["dias_incapacidad"], fmt="0", align="center"); col+=1
        _cell(ws, row, col, inc["dias_trabajados"] - inc["dias_incapacidad"], fmt="0", align="center"); col+=1
        _cell(ws, row, col, t["sbc_dia"] / t.get("factor_integracion", 1.0493), fmt="#,##0.000000"); col+=1

        # Percepciones
        _cell(ws, row, col, r["sueldo_fiscal"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r.get("despensa", 0) or 0, fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r.get("asistencia", 0) or 0, fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r.get("puntualidad", 0) or 0, fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["vacaciones_fiscal"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["total_he_fiscal"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["prima_vac_fiscal"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r.get("compensacion", 0) or 0, fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["suma_fiscal"], fmt="#,##0.00", bold=True); col+=1

        # Deducciones
        _cell(ws, row, col, r["cuota_obrera"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, 0, fmt="#,##0.00"); col+=1  # desc alimentos
        _cell(ws, row, col, 0, fmt="#,##0.00"); col+=1  # impuesto (sep)
        _cell(ws, row, col, r["isr_calcula"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["sub_emp_acre"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["isr_neto"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["sub_emp_neto"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, r["infonavit"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, redondeo, fmt="#,##0.0000"); col+=1
        _cell(ws, row, col, r["suma_deduc"], fmt="#,##0.00", bold=True); col+=1
        _cell(ws, row, col, r["neto_fiscal"], fmt="#,##0.00", bold=True, bg=VERDE_CLR); col+=1
        _cell(ws, row, col, pagado, fmt="#,##0.00", bold=True); col+=1
        _cell(ws, row, col, 0, fmt="0", align="center"); col+=1

        # Exentos
        _cell(ws, row, col, r["exento_he"], fmt="#,##0.00"); col+=1
        _cell(ws, row, col, int(inc.get("horas_extras_fiscales", 0)), fmt="0", align="center"); col+=1
        _cell(ws, row, col, r["exento_prima_vac"], fmt="#,##0.00"); col+=1

        # Acumular totales
        totales["sueldo"]       += r["sueldo_fiscal"]
        totales["hrs_ext"]      += r["total_he_fiscal"]
        totales["vac"]          += r["vacaciones_fiscal"]
        totales["prima_vac"]    += r["prima_vac_fiscal"]
        totales["suma_remun"]   += r["suma_fiscal"]
        totales["cuota_obrera"] += r["cuota_obrera"]
        totales["isr_calcula"]  += r["isr_calcula"]
        totales["sub_emp"]      += r["sub_emp_acre"]
        totales["isr_neto"]     += r["isr_neto"]
        totales["infonavit"]    += r["infonavit"]
        totales["suma_deduc"]   += r["suma_deduc"]
        totales["neto"]         += r["neto_fiscal"]
        totales["pagado"]       += pagado
        num_trab += 1

        # Alternar color de fila
        if row % 2 == 0:
            for c2 in ws[row]:
                if not c2.fill.fgColor.rgb or c2.fill.fgColor.rgb in ("00000000", VERDE_CLR):
                    c2.fill = PatternFill("solid", fgColor=GRIS_CLR)

        row += 1

    # Fila de totales
    ws.merge_cells(f"A{row}:O{row}")
    c = ws.cell(row=row, column=1, value=f"TOTAL  —  {num_trab} TRABAJADORES")
    c.font = Font(bold=True, size=9, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")

    for col, val in [
        (16, totales["sueldo"]), (20, totales["vac"]), (21, totales["hrs_ext"]),
        (22, totales["prima_vac"]), (24, totales["suma_remun"]),
        (25, totales["cuota_obrera"]), (28, totales["isr_calcula"]),
        (29, totales["sub_emp"]), (30, totales["isr_neto"]),
        (32, totales["infonavit"]), (34, totales["suma_deduc"]),
        (35, totales["neto"]), (36, totales["pagado"]),
    ]:
        c2 = ws.cell(row=row, column=col, value=val)
        c2.font = Font(bold=True, size=8, color=BLANCO)
        c2.fill = PatternFill("solid", fgColor=AZUL)
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.number_format = "#,##0.00"
        c2.border = _border
    ws.row_dimensions[row].height = 16

    # Freeze panes
    ws.freeze_panes = "A4"

    # ── Hoja 2: DIFERENCIAS (no fiscal) ────────────────────────────────
    ws2 = wb.create_sheet("DIFERENCIAS")

    ws2.merge_cells("A1:K1")
    c = ws2["A1"]
    c.value = f"DIFERENCIAS (PAGO NO FISCAL EN EFECTIVO) | SEMANA {periodo['num_semana']} / {periodo['anio']}"
    c.font = Font(bold=True, size=10, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=ROJO)
    c.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 18

    dif_headers = [
        ("NSS", 14), ("NOMBRE", 30), ("TIPO", 8), ("ÁREA", 20),
        ("NETO FISCAL", 12), ("NETO REAL", 12), ("DIFERENCIA\nEFECTIVO", 12),
        ("FORMA\nPAGO", 10), ("BANCO", 12), ("CUENTA/TARJETA", 20), ("OBSERVACIONES", 25),
    ]
    for col_idx, (hdr_name, col_w) in enumerate(dif_headers, start=1):
        _hdr(ws2, 2, col_idx, hdr_name, bg="FFCCCC", bold=True)
        ws2.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws2.row_dimensions[2].height = 28

    row2 = 3
    total_dif = 0.0
    for inc in incidencias_list:
        t = inc["trabajador"]
        r = inc["resultado"]
        if not r or r["diferencia"] <= 0:
            continue

        col = 1
        _cell(ws2, row2, col, t["imss"], align="left"); col+=1
        _cell(ws2, row2, col, t["nombre_completo"], align="left"); col+=1
        _cell(ws2, row2, col, t["tipo_trabajador"][0], align="center"); col+=1
        _cell(ws2, row2, col, t.get("area_funcional",""), align="left"); col+=1
        _cell(ws2, row2, col, r["neto_fiscal"], fmt="#,##0.00"); col+=1
        _cell(ws2, row2, col, r["neto_real"], fmt="#,##0.00"); col+=1
        _cell(ws2, row2, col, r["diferencia"], fmt="#,##0.00", bold=True, bg="FFCCCC"); col+=1
        _cell(ws2, row2, col, "EFECTIVO", align="center"); col+=1
        _cell(ws2, row2, col, t.get("banco",""), align="left"); col+=1
        _cell(ws2, row2, col, t.get("num_tarjeta", t.get("num_cuenta","")), align="left"); col+=1
        _cell(ws2, row2, col, inc.get("observacion",""), align="left"); col+=1

        total_dif += r["diferencia"]
        row2 += 1

    # Total diferencias
    ws2.merge_cells(f"A{row2}:F{row2}")
    c = ws2.cell(row=row2, column=1, value="TOTAL EFECTIVO A PAGAR")
    c.font = Font(bold=True, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=ROJO)
    c.alignment = Alignment(horizontal="center")
    c2 = ws2.cell(row=row2, column=7, value=total_dif)
    c2.font = Font(bold=True, color=BLANCO)
    c2.fill = PatternFill("solid", fgColor=ROJO)
    c2.number_format = "#,##0.00"

    ws2.freeze_panes = "A3"

    # ── Hoja 3: RESUMEN ────────────────────────────────────────────────
    ws3 = wb.create_sheet("RESUMEN")

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18

    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value = f"RESUMEN DE NÓMINA — SEMANA {periodo['num_semana']} / {periodo['anio']}"
    c.font = Font(bold=True, size=12, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22

    _hdr(ws3, 2, 1, "CONCEPTO", bold=True)
    _hdr(ws3, 2, 2, "PERMANENTE", bold=True)
    _hdr(ws3, 2, 3, "EVENTUAL", bold=True)
    _hdr(ws3, 2, 4, "TOTAL", bold=True)

    # Calcular resumen por tipo
    resumen = {"PERMANENTE": {}, "EVENTUAL": {}}
    conceptos_keys = [
        ("Trabajadores", "num"),
        ("Sueldo", "sueldo_fiscal"),
        ("Horas Extras (fiscal)", "total_he_fiscal"),
        ("Vacaciones (fiscal)", "vacaciones_fiscal"),
        ("Prima Vacacional (fiscal)", "prima_vac_fiscal"),
        ("TOTAL PERCEPCIONES", "suma_fiscal"),
        ("Cuota Obrera IMSS", "cuota_obrera"),
        ("ISR Calculado", "isr_calcula"),
        ("Subsidio al Empleo", "sub_emp_acre"),
        ("ISR Neto", "isr_neto"),
        ("Crédito INFONAVIT", "infonavit"),
        ("TOTAL DEDUCCIONES", "suma_deduc"),
        ("NETO FISCAL (transferencia)", "neto_fiscal"),
        ("DIFERENCIA (efectivo)", "diferencia"),
        ("NETO REAL (total)", "neto_real"),
    ]

    for tipo in ["PERMANENTE", "EVENTUAL"]:
        resumen[tipo] = {k[1]: 0.0 for k in conceptos_keys}

    for inc in incidencias_list:
        t = inc["trabajador"]
        r = inc["resultado"]
        if not r:
            continue
        tipo = t["tipo_trabajador"]
        if tipo not in resumen:
            continue
        resumen[tipo]["num"] += 1
        for _, key in conceptos_keys[1:]:
            resumen[tipo][key] = resumen[tipo].get(key, 0.0) + r.get(key, 0.0)

    separadores = {"TOTAL PERCEPCIONES", "TOTAL DEDUCCIONES", "NETO FISCAL (transferencia)", "NETO REAL (total)"}
    row3 = 3
    for label, key in conceptos_keys:
        is_sep = label in separadores
        bg = AZUL_CLR if is_sep else None
        fmt = "#,##0" if key == "num" else "#,##0.00"
        _cell(ws3, row3, 1, label, align="left", bold=is_sep, bg=bg)
        vp = resumen["PERMANENTE"].get(key, 0.0)
        ve = resumen["EVENTUAL"].get(key, 0.0)
        vt = vp + ve
        _cell(ws3, row3, 2, vp, fmt=fmt, bold=is_sep, bg=bg)
        _cell(ws3, row3, 3, ve, fmt=fmt, bold=is_sep, bg=bg)
        _cell(ws3, row3, 4, vt, fmt=fmt, bold=is_sep, bg=bg)
        row3 += 1

    ws3.freeze_panes = "A3"

    # ── Guardar ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
